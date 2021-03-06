# excel.py
# This is various practice on working with excel-style spreadsheets

import openpyxl
from openpyxl.utils import get_column_letter


def main():
    ''' The main function of the program. '''

    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb['Sheet1']

    # print(sheet.max_row)
    # print(sheet.max_column)
    print(get_column_letter(sheet.max_column))

    for i in range(1, sheet.max_row):
        print(i, sheet.cell(row=i, column=2).value,
              sheet.cell(row=i, column=3).value)

    rows = tuple(sheet['A1':'C6'])

    for row_objects in rows:
        for cell_object in row_objects:
            print(cell_object.coordinate, cell_object.value)
        print(f'<---END OF ROW--->\n')

    wb.close()


if __name__ == '__main__':
    main()
